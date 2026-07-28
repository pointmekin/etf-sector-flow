import hashlib
from dataclasses import dataclass
from datetime import UTC, date, datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from tempfile import NamedTemporaryFile

import httpx
from openpyxl import load_workbook

EXPECTED_HEADERS = ("Date", "NAV", "Shares Outstanding", "Total Net Assets")


@dataclass(frozen=True)
class FundDailyInput:
    date: date
    nav: Decimal | None
    shares_outstanding: Decimal | None
    aum: Decimal | None
    source_url: str
    source_hash: str
    retrieved_at: datetime
    quality_status: str = "ok"
    quality_note: str | None = None


@dataclass(frozen=True)
class PriceInput:
    date: date
    adjusted_close: Decimal


async def download_to_temp(client: httpx.AsyncClient, url: str) -> Path:
    with NamedTemporaryFile(suffix=".xlsx", delete=False) as output:
        async with client.stream("GET", url, follow_redirects=True) as response:
            response.raise_for_status()
            async for chunk in response.aiter_bytes():
                output.write(chunk)
        return Path(output.name)


def parse_nav_history(path: Path, source_url: str) -> list[FundDailyInput]:
    source_hash = hashlib.sha256(path.read_bytes()).hexdigest()
    retrieved_at = datetime.now(UTC)
    sheet = load_workbook(path, read_only=True, data_only=True).active
    headers = tuple(sheet.cell(4, column).value for column in range(1, 5))
    if headers != EXPECTED_HEADERS:
        raise ValueError(f"Unexpected State Street columns: {headers}")

    rows: list[FundDailyInput] = []
    seen_dates: set[date] = set()
    for values in sheet.iter_rows(min_row=5, max_col=4, values_only=True):
        if not values[0]:
            continue
        try:
            parsed_date = _parse_date(values[0])
        except ValueError:
            if rows:
                break
            raise ValueError(f"Invalid source date: {values[0]}") from None
        if parsed_date in seen_dates:
            raise ValueError(f"Duplicate source date: {parsed_date}")
        seen_dates.add(parsed_date)
        nav, shares, aum = (_decimal_or_none(value) for value in values[1:4])
        numeric_values = [value for value in (nav, shares, aum) if value is not None]
        if any(value < 0 for value in numeric_values):
            raise ValueError(f"Negative source value for {parsed_date}")

        if nav is None or shares is None or aum is None:
            quality_status = "unavailable"
            quality_note = "NAV, shares, or AUM is unavailable in the source"
        else:
            mismatch = abs(aum - nav * shares) / aum if aum else Decimal(0)
            quality_status = "review" if mismatch > Decimal("0.02") else "ok"
            quality_note = (
                "AUM differs from NAV × shares by more than 2%"
                if quality_status == "review"
                else None
            )
        rows.append(
            FundDailyInput(
                date=parsed_date,
                nav=nav,
                shares_outstanding=shares,
                aum=aum,
                source_url=source_url,
                source_hash=source_hash,
                retrieved_at=retrieved_at,
                quality_status=quality_status,
                quality_note=quality_note,
            )
        )
    if not rows:
        raise ValueError("State Street workbook contains no data rows")
    return sorted(rows, key=lambda row: row.date)


async def fetch_adjusted_prices(
    client: httpx.AsyncClient, ticker: str, api_key: str
) -> list[PriceInput]:
    response = await client.get(
        "https://api.twelvedata.com/time_series",
        params={
            "symbol": ticker,
            "interval": "1day",
            "outputsize": 5000,
            "adjust": "all",
            "order": "asc",
            "apikey": api_key,
        },
    )
    response.raise_for_status()
    payload = response.json()
    if payload.get("status") == "error":
        raise ValueError(
            f"Twelve Data error for {ticker}: {payload.get('message', 'unknown error')}"
        )
    return [
        PriceInput(
            date=date.fromisoformat(item["datetime"][:10]),
            adjusted_close=Decimal(item["close"]),
        )
        for item in payload.get("values", [])
    ]


def _parse_date(value: object) -> date:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    return datetime.strptime(str(value), "%d-%b-%Y").date()


def _decimal_or_none(value: object) -> Decimal | None:
    if value in (None, "", "-"):
        return None
    try:
        return Decimal(str(value))
    except (InvalidOperation, TypeError) as error:
        raise ValueError(f"Invalid numeric value: {value}") from error
